import csv
import io
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from jinja2 import Environment, FileSystemLoader
from ..db import fetchone, fetchall


TEMPLATE_DIR = Path(__file__).parent.parent / "templates"


def _load_quote_data(quote_id: int, db: sqlite3.Connection) -> dict:
    quote = fetchone(
        db,
        """
        SELECT q.*, cl.name as client_name, co.name as country_name,
               cu.code as currency_code, u.name as created_by_name
        FROM quotes q
        JOIN clients cl ON cl.id = q.client_id
        JOIN countries co ON co.id = q.country_id
        JOIN currencies cu ON cu.id = q.currency_id
        JOIN users u ON u.id = q.created_by
        WHERE q.id = ?
        """,
        (quote_id,),
    )
    if not quote:
        raise ValueError(f"Quote {quote_id} not found")

    items = fetchall(
        db,
        """
        SELECT li.*, p.name as product_name, q2.name as quality_name,
               ps.display_name as package_name
        FROM quote_line_items li
        JOIN products p ON p.id = li.product_id
        JOIN qualities q2 ON q2.id = li.quality_id
        JOIN package_sizes ps ON ps.id = li.package_size_id
        WHERE li.quote_id = ?
        """,
        (quote_id,),
    )

    settings = {
        r["key"]: r["value"]
        for r in fetchall(
            db,
            "SELECT key, value FROM system_settings WHERE key IN ('pdf_company_name','pdf_company_address','pdf_terms')",
        )
    }

    return {
        "quote": dict(quote),
        "items": [dict(i) for i in items],
        "settings": settings,
    }


def export_pdf(quote_id: int, db: sqlite3.Connection) -> bytes:
    from weasyprint import HTML
    data = _load_quote_data(quote_id, db)
    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template("exports/quote_pdf.html")
    html_content = template.render(**data)
    return HTML(string=html_content).write_pdf()


def export_xlsx(quote_id: int, db: sqlite3.Connection) -> bytes:
    data = _load_quote_data(quote_id, db)
    quote = data["quote"]
    items = data["items"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    ws["A1"] = data["settings"].get("pdf_company_name", "S&G Exports")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Quote: {quote['quote_number']}  |  Client: {quote['client_name']}  |  Date: {quote['created_at'][:10]}"

    headers = [
        "Product", "Quality", "Package", "Qty (kg)", "Vendor Price/kg",
        "Cost/kg (INR)", "Margin %", "Selling Price/kg (INR)", f"Price ({quote['currency_code']})",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(items, 5):
        ws.cell(row=row_idx, column=1, value=item["product_name"])
        ws.cell(row=row_idx, column=2, value=item["quality_name"])
        ws.cell(row=row_idx, column=3, value=item["package_name"])
        ws.cell(row=row_idx, column=4, value=item["quantity_kg"])
        ws.cell(row=row_idx, column=5, value=item["vendor_price_per_kg"])
        ws.cell(row=row_idx, column=6, value=item["cost_per_kg_inr"])
        ws.cell(row=row_idx, column=7, value=item["margin_pct"])
        ws.cell(row=row_idx, column=8, value=item["selling_price_per_kg_inr"])
        ws.cell(row=row_idx, column=9, value=item.get("selling_price_fx"))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv(quote_id: int, db: sqlite3.Connection) -> str:
    data = _load_quote_data(quote_id, db)
    quote = data["quote"]
    items = data["items"]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Quote Number", "Client", "Country", "Currency",
        "Product", "Quality", "Package", "Qty (kg)",
        "Vendor Price/kg", "Cost/kg INR", "Margin %",
        "Selling Price/kg INR", f"Price {quote['currency_code']}",
    ])
    for item in items:
        writer.writerow([
            quote["quote_number"], quote["client_name"], quote["country_name"], quote["currency_code"],
            item["product_name"], item["quality_name"], item["package_name"], item["quantity_kg"],
            item["vendor_price_per_kg"], item["cost_per_kg_inr"], item["margin_pct"],
            item["selling_price_per_kg_inr"], item.get("selling_price_fx"),
        ])
    return output.getvalue()


def export_analytics_xlsx(rows: list[dict], columns: list[str]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analytics"
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c) for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
