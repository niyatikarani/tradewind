#!/usr/bin/env python3
"""Seed quotation.db with comprehensive demo data from Product Library.xlsx"""
import sqlite3
import sys
import argparse
import json
import getpass
from pathlib import Path
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    sys.exit("Install openpyxl: pip install openpyxl")

try:
    import bcrypt
except ImportError:
    sys.exit("Install bcrypt: pip install bcrypt")

SCHEMA_PATH = Path(__file__).parent / "schema.sql"

def run_seed(db_path: str, product_lib: Path) -> None:
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA journal_mode=WAL")
    conn.row_factory = sqlite3.Row

    # Create schema first (idempotent — uses IF NOT EXISTS)
    conn.executescript(SCHEMA_PATH.read_text())
    conn.commit()

    # Clean all rows so re-running seed always produces a clean state
    conn.executescript("PRAGMA foreign_keys=OFF")
    for table in ["rule_change_log", "field_override_log", "quote_line_items", "quotes",
                  "client_price_locks", "country_product_overrides", "user_features",
                  "clients", "countries", "list_prices", "fx_rates", "vendor_prices",
                  "certification_types", "sanitization_costs", "cbm_dimensions",
                  "packaging_materials", "labour_rates", "package_sizes", "qualities",
                  "products", "system_settings", "margin_factors", "feature_templates",
                  "currencies", "features", "users"]:
        conn.execute(f"DELETE FROM {table}")
    conn.executescript("PRAGMA foreign_keys=ON")
    conn.commit()

    print("\n=== SEEDING DATABASE WITH COMPREHENSIVE DEMO DATA ===\n")

    # FEATURES & RBAC
    print("[1/11] Seeding features...")
    features = [
        ("price.entry", "Price Entry", "Enter vendor prices", "prices"),
        ("price.bulk", "Bulk Price Entry", "Enter prices in bulk", "prices"),
        ("price.stale", "Stale Dashboard", "View stale price alerts", "prices"),
        ("price.trends", "Price Trends", "View price trend charts", "prices"),
        ("quote.create", "Create Quotes", "Create and manage quotes", "quotes"),
        ("quote.manual_override", "Manual Override", "Override calculated fields", "quotes"),
        ("quote.stale_override", "Stale Override", "Override stale price blocks", "quotes"),
        ("quote.export", "Export Quotes", "Export quotes to PDF/XLSX/CSV", "quotes"),
        ("rule_engine.view", "View Rules", "View business rules", "rule_engine"),
        ("rule_engine.edit", "Edit Rules", "Edit business rules", "rule_engine"),
        ("analytics.view", "View Analytics", "View analytics dashboards", "analytics"),
        ("admin.master_data", "Master Data Admin", "Manage products, sizes, etc.", "admin"),
        ("admin.users", "User Admin", "Manage users and permissions", "admin"),
        ("admin.system", "System Settings", "Manage system configuration", "admin"),
    ]
    for key, label, desc, cat in features:
        conn.execute(
            "INSERT OR IGNORE INTO features(key, label, description, category) VALUES(?,?,?,?)",
            (key, label, desc, cat)
        )
    conn.commit()
    print(f"  Inserted {len(features)} features")

    # CURRENCIES
    print("[2/11] Seeding currencies...")
    currencies = [
        ("USD", "US Dollar", 0.02),
        ("INR", "Indian Rupee", 0.0),
        ("EUR", "Euro", 0.02),
        ("GBP", "British Pound", 0.02),
        ("AED", "UAE Dirham", 0.01),
        ("CAD", "Canadian Dollar", 0.02),
        ("AUD", "Australian Dollar", 0.02),
    ]
    for code, name, buf in currencies:
        conn.execute(
            "INSERT OR IGNORE INTO currencies(code, name, fx_variation_buffer) VALUES(?,?,?)",
            (code, name, buf)
        )
    conn.commit()
    print(f"  Inserted {len(currencies)} currencies")

    # SYSTEM SETTINGS
    print("[3/11] Seeding system settings...")
    defaults = {
        "stale_days": "7",
        "base_margin": "10",
        "margin_floor": "3",
        "margin_cap": "25",
        "transport_per_kg": "1.50",
        "loading_per_kg": "0.80",
        "cha_flat": "25000",
        "customs_flat": "15000",
        "container_20ft_cbm": "28",
        "container_20ft_kg": "21700",
        "container_40ft_cbm": "56",
        "container_40ft_kg": "26500",
        "quote_number_prefix": "SGE",
        "quote_number_next": "1001",
    }
    for key, value in defaults.items():
        conn.execute(
            "INSERT OR REPLACE INTO system_settings(key, value) VALUES(?,?)",
            (key, value)
        )
    conn.commit()
    print(f"  Inserted {len(defaults)} system settings")

    # MARGIN FACTORS
    print("[4/11] Seeding margin factors...")
    factors = [
        ("volume", "Order Volume (kg)", 1.5, [{"max": 500, "adjustment": -1.0}, {"min": 500, "max": 2000, "adjustment": 0}, {"min": 2000, "adjustment": 1.5}]),
        ("advance_payment", "Advance Payment %", 1.2, [{"max": 25, "adjustment": -1.0}, {"min": 25, "max": 50, "adjustment": 0}, {"min": 50, "adjustment": 1.0}]),
        ("client_history", "Client History", 1.0, [{"max": 1, "adjustment": -0.5}, {"min": 1, "max": 5, "adjustment": 0}, {"min": 5, "adjustment": 1.0}]),
        ("competition", "Competition Level", 1.0, [{"max": 1, "adjustment": 1.5}, {"min": 1, "max": 2, "adjustment": 0}, {"min": 2, "adjustment": -1.0}]),
        ("price_volatility", "Price Volatility", 0.8, [{"max": 1, "adjustment": 0}, {"min": 1, "max": 3, "adjustment": -0.5}, {"min": 3, "adjustment": -1.0}]),
        ("country_risk", "Country Risk Score", 1.0, [{"max": 2, "adjustment": -0.5}, {"min": 2, "max": 4, "adjustment": 0}, {"min": 4, "adjustment": 1.5}]),
        ("urgency", "Order Urgency", 0.8, [{"max": 1, "adjustment": -0.5}, {"min": 1, "max": 2, "adjustment": 0}, {"min": 2, "adjustment": 1.0}]),
    ]
    for key, label, weight, tiers in factors:
        conn.execute(
            "INSERT OR IGNORE INTO margin_factors(factor_key, label, weight, scoring_tiers) VALUES(?,?,?,?)",
            (key, label, weight, json.dumps(tiers))
        )
    conn.commit()
    print(f"  Inserted {len(factors)} margin factors")

    # PRODUCTS & QUALITIES from Product Library.xlsx
    print("[5/11] Seeding products & qualities from Product Library.xlsx...")
    product_count = 0
    quality_count = 0

    try:
        # Try parent directory first
        parent_lib = Path(__file__).parent.parent / "Product Library.xlsx"
        if parent_lib.exists():
            product_lib = parent_lib

        if product_lib.exists():
            wb = openpyxl.load_workbook(product_lib, data_only=True)

            # Load products
            if "Product" in wb.sheetnames:
                ws = wb["Product"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    prod_id, prod_name = row[1], row[2]
                    if prod_id and prod_name:
                        prod_name_clean = str(prod_name).strip()
                        competition = "high" if "rice" in prod_name_clean.lower() else "medium"
                        conn.execute(
                            "INSERT OR IGNORE INTO products(name, category, competition_level) VALUES(?,?,?)",
                            (prod_name_clean, "Spices", competition)
                        )
                        product_count += 1

            # Load qualities
            if "Quality" in wb.sheetnames:
                ws = wb["Quality"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    prod_id, qual_id, qual_name = row[0], row[1], row[2]
                    if prod_id and qual_name:
                        prod_id_clean = str(prod_id).strip()
                        qual_name_clean = str(qual_name).strip()

                        prod = conn.execute(
                            "SELECT id FROM products WHERE category='Spices' LIMIT 1 OFFSET ?",
                            (quality_count % max(1, product_count - 1),)
                        ).fetchone()

                        if prod:
                            conn.execute(
                                "INSERT OR IGNORE INTO qualities(product_id, name, sort_order) VALUES(?,?,?)",
                                (prod["id"], qual_name_clean, quality_count % 5)
                            )
                            quality_count += 1

            conn.commit()
            print(f"  Loaded {product_count} products, {quality_count} qualities")
        else:
            print(f"  SKIP: {product_lib} not found")
    except Exception as e:
        print(f"  SKIP: Error reading Product Library: {e}")

    # PACKAGE SIZES
    print("[6/11] Seeding package sizes...")
    sizes = [
        (20, "20g"), (50, "50g"), (100, "100g"), (200, "200g"), (250, "250g"),
        (500, "500g"), (1000, "1kg"), (2000, "2kg"), (5000, "5kg"),
        (10000, "10kg"), (15000, "15kg"), (20000, "20kg"), (25000, "25kg"),
    ]
    for weight_g, display in sizes:
        conn.execute(
            "INSERT OR IGNORE INTO package_sizes(display_name, weight_grams, is_standard) VALUES(?,?,?)",
            (display, weight_g, 1)
        )

        pkg = conn.execute("SELECT id FROM package_sizes WHERE display_name=?", (display,)).fetchone()
        pkg_id = pkg["id"]

        # Labour rates
        conn.execute(
            "INSERT OR REPLACE INTO labour_rates(package_size_id, packing_cost, sticker_1side, sticker_2side) VALUES(?,?,?,?)",
            (pkg_id, 10.0, 2.0, 3.0)
        )

        # Packaging materials
        conn.execute(
            "INSERT OR REPLACE INTO packaging_materials(package_size_id, pkts_per_carton, carton_rate, sticker_rate, pouch_price) VALUES(?,?,?,?,?)",
            (pkg_id, 50, 25.0, 0.5, 2.0)
        )

        # CBM dimensions
        conn.execute(
            "INSERT OR REPLACE INTO cbm_dimensions(package_size_id, length_cm, breadth_cm, height_cm) VALUES(?,?,?,?)",
            (pkg_id, 10.0, 8.0, 12.0)
        )

    conn.commit()
    print(f"  Seeded {len(sizes)} package sizes")

    # COUNTRIES
    print("[7/11] Seeding countries...")
    countries = [
        ("USA", "US", 1, 1, "Standard FDA"),
        ("UAE", "AE", 2, 2, "Halal preferred"),
        ("UK", "GB", 1, 1, "EU standards"),
        ("EU", "EU", 1, 1, "Strict safety"),
        ("Canada", "CA", 1, 1, "CFIA reqs"),
        ("Australia", "AU", 1, 2, "Stringent"),
        ("Singapore", "SG", 2, 2, "High quality"),
        ("China", "CN", 3, 3, "Payment risk"),
        ("Japan", "JP", 1, 1, "Premium"),
        ("India", "IN", 1, 1, "Domestic"),
    ]
    for name, code, risk, _, notes in countries:
        inr_curr = conn.execute("SELECT id FROM currencies WHERE code='INR'").fetchone()
        conn.execute(
            "INSERT OR IGNORE INTO countries(name, code, default_currency_id, risk_score, notes) VALUES(?,?,?,?,?)",
            (name, code, inr_curr["id"] if inr_curr else 2, risk, notes)
        )
    conn.commit()
    print(f"  Seeded {len(countries)} countries")

    # SANITIZATION & CERTIFICATIONS
    print("[8/11] Seeding sanitization & certifications...")
    products = conn.execute("SELECT id FROM products LIMIT 5").fetchall()
    for prod in products:
        for san_type, cost in [("steam", 100.0), ("chemical", 150.0), ("none", 0.0)]:
            conn.execute(
                "INSERT OR IGNORE INTO sanitization_costs(product_id, type, cost_per_kg) VALUES(?,?,?)",
                (prod["id"], san_type, cost)
            )

    certs = [("FDA", 250.0), ("Spice Board", 150.0), ("ISO 22000", 300.0), ("Halal", 200.0), ("Organic", 350.0)]
    for cert_name, cost in certs:
        conn.execute(
            "INSERT OR IGNORE INTO certification_types(name, cost_per_kg) VALUES(?,?)",
            (cert_name, cost)
        )
    conn.commit()
    print(f"  Seeded sanitization types & {len(certs)} certifications")

    # USERS
    print("[9/11] Seeding demo users...")
    demo_pwd = bcrypt.hashpw(b"demo123456", bcrypt.gensalt()).decode()
    users = [
        ("Ravi Sharma", "ravi@sandgexports.com", demo_pwd),
        ("Priya Patel", "priya@sandgexports.com", demo_pwd),
        ("Suresh Kumar", "suresh@sandgexports.com", demo_pwd),
        ("Marketing Team", "marketing@sandgexports.com", demo_pwd),
    ]
    user_map = {}
    for name, email, pwd in users:
        conn.execute(
            "INSERT OR IGNORE INTO users(name, email, password_hash) VALUES(?,?,?)",
            (name, email, pwd)
        )
        user = conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone()
        user_map[email] = user["id"]

    conn.commit()

    # Assign features to demo users
    role_features = {
        "ravi@sandgexports.com":      ["price.entry", "price.bulk", "price.stale", "price.trends"],
        "priya@sandgexports.com":     ["quote.create", "quote.export", "price.stale", "price.trends"],
        "suresh@sandgexports.com":    ["price.entry", "price.bulk", "price.stale", "price.trends",
                                       "quote.create", "quote.manual_override", "quote.stale_override",
                                       "quote.export", "analytics.view", "rule_engine.view"],
        "marketing@sandgexports.com": ["quote.create", "quote.export"],
    }
    for email, feature_keys in role_features.items():
        uid = user_map[email]
        for fkey in feature_keys:
            feat = conn.execute("SELECT id FROM features WHERE key=?", (fkey,)).fetchone()
            if feat:
                conn.execute(
                    "INSERT OR IGNORE INTO user_features(user_id, feature_id) VALUES(?,?)",
                    (uid, feat["id"])
                )
    conn.commit()
    print(f"  Seeded {len(users)} demo users")

    # FX RATES  (USD≈83, EUR≈90, GBP≈105, AED≈22.6, CAD≈61, AUD≈54)
    print("[10a/11] Seeding FX rates...")
    fx_rates = [
        ("USD", 83.50),
        ("EUR", 90.20),
        ("GBP", 104.80),
        ("AED", 22.72),
        ("CAD", 61.30),
        ("AUD", 54.40),
    ]
    ravi_id = user_map.get("ravi@sandgexports.com", 1)
    for cur_code, rate in fx_rates:
        cur = conn.execute("SELECT id FROM currencies WHERE code=?", (cur_code,)).fetchone()
        if cur:
            conn.execute(
                "INSERT INTO fx_rates(currency_id, rate_vs_inr, source, entered_by) VALUES(?,?,?,?)",
                (cur["id"], rate, "manual_seed", ravi_id)
            )
    conn.commit()
    print(f"  Seeded {len(fx_rates)} FX rates")

    # VENDOR PRICES
    print("[10/11] Seeding vendor prices...")
    qualities = conn.execute("SELECT id, product_id FROM qualities LIMIT 40").fetchall()
    vendors = ["Sharma Spices", "Patel & Co", "Global Trade", "Mumbai Traders"]
    prices_added = 0
    admin_user_id = user_map.get("ravi@sandgexports.com", 1)

    for quality in qualities:
        for i, vendor in enumerate(vendors):
            price = 100 + (quality["id"] % 500) + (i * 5)
            conn.execute(
                "INSERT INTO vendor_prices(product_id, quality_id, price_per_kg, vendor_name, entered_by) VALUES(?,?,?,?,?)",
                (quality["product_id"], quality["id"], price, vendor, admin_user_id)
            )
            prices_added += 1

    conn.commit()
    print(f"  Seeded {prices_added} vendor prices")

    # CLIENTS
    print("[11/11] Seeding sample clients...")
    client_countries = [
        ("Premium Foods Inc", "US"),
        ("Spice World Trading", "AE"),
        ("London Imports Ltd", "GB"),
        ("Berlin GmbH", "EU"),
        ("Toronto Distributor", "CA"),
        ("Sydney Wholesale", "AU"),
        ("Singapore Retail", "SG"),
    ]
    for client_name, country_code in client_countries:
        country = conn.execute("SELECT id FROM countries WHERE code=?", (country_code,)).fetchone()
        if country:
            conn.execute(
                "INSERT INTO clients(name, country_id, payment_risk, credit_terms) VALUES(?,?,?,?)",
                (client_name, country["id"], "low" if "Premium" in client_name else "medium", "Net 30" if "Premium" in client_name else "Net 60")
            )

    # SAMPLE QUOTES
    quote_num = 1001
    clients = conn.execute("SELECT id, country_id FROM clients").fetchall()
    products_list = conn.execute("SELECT p.id, q.id as quality_id FROM products p JOIN qualities q ON p.id=q.product_id LIMIT 10").fetchall()
    pkg_sizes = conn.execute("SELECT id FROM package_sizes LIMIT 3").fetchall()
    usd_curr = conn.execute("SELECT id FROM currencies WHERE code='USD'").fetchone()

    for client in clients[:3]:
        quote_number = f"SGE-{quote_num}"
        conn.execute(
            "INSERT INTO quotes(quote_number, client_id, country_id, currency_id, status, created_by) VALUES(?,?,?,?,?,?)",
            (quote_number, client["id"], client["country_id"], usd_curr["id"], "draft", admin_user_id)
        )
        quote_num += 1

        quote = conn.execute("SELECT id FROM quotes WHERE quote_number=?", (quote_number,)).fetchone()
        for prod in products_list[:3]:
            pkg = pkg_sizes[0]
            conn.execute(
                """INSERT INTO quote_line_items(
                    quote_id, product_id, quality_id, package_size_id,
                    quantity_kg, sanitization_type, label_sides, certifications,
                    vendor_price_per_kg, raw_material_inr, labour_inr,
                    packaging_inr, sanitization_inr, certification_inr,
                    loading_inr, transport_inr, subtotal_inr, cost_per_kg_inr,
                    margin_pct, selling_price_per_kg_inr, selling_price_fx,
                    fx_rate_used, cbm_per_carton, num_cartons, total_cbm
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (quote["id"], prod["id"], prod["quality_id"], pkg["id"],
                 250, "steam", 1, "[]",
                 150.0, 37500, 500,
                 300, 200, 100,
                 200, 300, 39000, 156,
                 10, 166, 13695,
                 82.5, 0.5, 3, 1.5)
            )

    # Update quote_number_next so new quotes don't collide with seeded ones
    conn.execute("UPDATE system_settings SET value=? WHERE key='quote_number_next'", (str(quote_num),))
    conn.commit()
    print(f"  Seeded {len(client_countries)} clients with sample quotes (next quote# SGE-{quote_num})")

    print("\n[SETUP] Creating admin user...")
    admin_pwd = getpass.getpass("  Admin password: ")
    confirm = getpass.getpass("  Confirm:  ")
    if admin_pwd != confirm:
        sys.exit("Passwords do not match")

    admin_hash = bcrypt.hashpw(admin_pwd.encode(), bcrypt.gensalt()).decode()
    conn.execute(
        "INSERT INTO users(name, email, password_hash) VALUES(?,?,?)",
        ("Admin", "admin@sandgexports.com", admin_hash)
    )
    admin_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # Assign all features to admin
    for feature in conn.execute("SELECT id FROM features").fetchall():
        conn.execute(
            "INSERT OR IGNORE INTO user_features(user_id, feature_id) VALUES(?,?)",
            (admin_id, feature["id"])
        )

    conn.commit()
    print(f"  Created admin user")

    conn.close()
    print("\n[SUCCESS] DATABASE SEED COMPLETE!")
    print("\nLogin credentials:")
    print(f"  admin@sandgexports.com (password set above)")
    print(f"  ravi@sandgexports.com / demo123456")
    print(f"  priya@sandgexports.com / demo123456")
    print(f"  suresh@sandgexports.com / demo123456")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Seed quotation.db")
    parser.add_argument("--db", default="quotation.db")
    parser.add_argument("--product-lib", default=None)
    args = parser.parse_args()

    product_lib = args.product_lib
    if not product_lib:
        parent_path = Path(__file__).parent.parent / "Product Library.xlsx"
        product_lib = str(parent_path if parent_path.exists() else "Product Library.xlsx")

    run_seed(args.db, Path(product_lib))
